"""Bootstrap del primo template PN a partire dal ciclo reale fornito.

Tokenizza il ciclo compilato (sostituisce i valori variabili con i token
``{{...}}``) e produce:
    templates/70009309-P002/template.xlsx
    templates/70009309-P002/config.json

Questo script e' il "seme" del futuro Creatore di template (Fase 3):
la logica di mappatura celle -> token qui e' scritta a mano; il wizard la
rendera' interattiva/automatica.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple

QUI = Path(__file__).resolve().parent
RADICE = QUI.parent
SORGENTE = Path(r"C:/Users/Lorenzo/Desktop/6935-26_ODL_OR_70009309-P002_REV.1.xlsx")
PN = "70009309-P002"
DEST_DIR = RADICE / "templates" / PN


def anchor_merge(ws, coord: str) -> str:
    """Se ``coord`` cade dentro una cella unita, ritorna la cella in alto a sx."""
    r, c = coordinate_to_tuple(coord)
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(mr))
        if min_r <= r <= max_r and min_c <= c <= max_c:
            return ws.cell(row=min_r, column=min_c).coordinate
    return coord


def set_cella(ws, coord: str, valore) -> None:
    ws[anchor_merge(ws, coord)].value = valore


def sostituisci_testo(ws, coord: str, vecchio: str, nuovo: str) -> None:
    ancora = anchor_merge(ws, coord)
    corrente = ws[ancora].value
    if corrente is None:
        ws[ancora].value = nuovo
        return
    ws[ancora].value = str(corrente).replace(vecchio, nuovo)


# Mappatura cella intera -> token
TOKEN_CELLE = {
    "C5": "{{COMMESSA}}",
    "M5": "{{ORDINE_ACQUISTO}}",
    "V5": "{{ORDINE_INTERNO}}",
    "E7": "{{NR_PEZZI}}",
    "M7": "{{DATA_ODA}}",
    "V7": "{{DATA_OI}}",
    "AF7": "{{DATA_CONSEGNA}}",
    "Y9": "{{REV}}",
    # Distinta base - materia prima / consumo
    "AD23": "{{LOTTO_MPL-076}}",
    "AD24": "{{LOTTO_3V000107}}",
    # Distinta base - inserti (lotti opzionali)
    "AD28": "{{LOTTO_5N005893}}",
    "AD29": "{{LOTTO_5N005231}}",
    "AD30": "{{LOTTO_5A007666}}",
    "AD31": "{{LOTTO_5A008049}}",
    "AD32": "{{LOTTO_5A008050}}",
    "AD33": "{{LOTTO_3K000005}}",
    "AD34": "{{LOTTO_3N000088}}",
}

# Sostituzioni testuali dentro le fasi (propagazione lotti)
SOSTITUZIONI_TESTO = [
    ("S40", "6985446", "{{LOTTO_ADESIVO_EPOX}}"),   # fase 20 - adesivo epossidico
    ("AA40", "Scadenza: ", "Scadenza: {{SCADENZA_ADESIVO_EPOX}}"),
    ("S45", "211", "{{LOTTO_3V000107}}"),            # fase 60 - stesso lotto Zucchini
]

# Definizione dei campi per il form (config.json)
CAMPI = [
    # Testata
    {"token": "{{COMMESSA}}", "etichetta": "Commessa", "tipo": "testo",
     "obbligatorio": True, "gruppo": "Testata"},
    {"token": "{{ORDINE_ACQUISTO}}", "etichetta": "Ordine d'Acquisto", "tipo": "testo",
     "gruppo": "Testata"},
    {"token": "{{ORDINE_INTERNO}}", "etichetta": "Ordine Interno (primo)",
     "tipo": "ordine_interno", "obbligatorio": True, "gruppo": "Testata",
     "aiuto": "Formato NNNN-YY, es. 6935-26. Si incrementa da solo per i cicli successivi."},
    {"token": "{{NR_PEZZI}}", "etichetta": "Nr. Pezzi", "tipo": "intero",
     "gruppo": "Testata"},
    {"token": "{{DATA_ODA}}", "etichetta": "Data ODA", "tipo": "data",
     "gruppo": "Testata", "aiuto": "gg/mm/aaaa"},
    {"token": "{{DATA_OI}}", "etichetta": "Data OI", "tipo": "data",
     "gruppo": "Testata", "aiuto": "gg/mm/aaaa"},
    {"token": "{{DATA_CONSEGNA}}", "etichetta": "Data Consegna", "tipo": "data",
     "gruppo": "Testata", "aiuto": "gg/mm/aaaa (opzionale)"},
    {"token": "{{REV}}", "etichetta": "Rev.", "tipo": "testo", "default": "1",
     "gruppo": "Testata"},
    # Materia prima / consumo
    {"token": "{{LOTTO_MPL-076}}", "etichetta": "MPL-076 - Divinycell H60",
     "tipo": "lotto", "gruppo": "Materia prima / consumo"},
    {"token": "{{LOTTO_3V000107}}", "etichetta": "3V000107 - Glue Zucchini 9631DM",
     "tipo": "lotto", "gruppo": "Materia prima / consumo"},
    {"token": "{{LOTTO_ADESIVO_EPOX}}", "etichetta": "Adesivo epossidico (fase 20)",
     "tipo": "lotto", "gruppo": "Materia prima / consumo"},
    {"token": "{{SCADENZA_ADESIVO_EPOX}}", "etichetta": "Scadenza adesivo epossidico",
     "tipo": "data", "gruppo": "Materia prima / consumo", "aiuto": "gg/mm/aaaa (opzionale)"},
    # Inserti (lotti opzionali)
    {"token": "{{LOTTO_5N005893}}", "etichetta": "5N005893 - Inserto Nylon",
     "tipo": "lotto", "gruppo": "Inserti (lotti, opzionali)"},
    {"token": "{{LOTTO_5N005231}}", "etichetta": "5N005231 - Inserto Nylon",
     "tipo": "lotto", "gruppo": "Inserti (lotti, opzionali)"},
    {"token": "{{LOTTO_5A007666}}", "etichetta": "5A007666 - Piastra alluminio",
     "tipo": "lotto", "gruppo": "Inserti (lotti, opzionali)"},
    {"token": "{{LOTTO_5A008049}}", "etichetta": "5A008049 - Piastra alluminio",
     "tipo": "lotto", "gruppo": "Inserti (lotti, opzionali)"},
    {"token": "{{LOTTO_5A008050}}", "etichetta": "5A008050 - Piastra alluminio",
     "tipo": "lotto", "gruppo": "Inserti (lotti, opzionali)"},
    {"token": "{{LOTTO_3K000005}}", "etichetta": "3K000005 - Rivetti",
     "tipo": "lotto", "gruppo": "Inserti (lotti, opzionali)"},
    {"token": "{{LOTTO_3N000088}}", "etichetta": "3N000088 - Cricchetto",
     "tipo": "lotto", "gruppo": "Inserti (lotti, opzionali)"},
]

CONFIG = {
    "pn": PN,
    "descrizione": "Assieme incollato copribracciolo CNT Sx",
    "template": "template.xlsx",
    "pattern_nome_file": "{ORDINE_INTERNO}_ODL_OR_{PN}_REV.{REV}.xlsx",
    "campi": CAMPI,
}


def main() -> None:
    DEST_DIR.mkdir(parents=True, exist_ok=True)
    # rich_text=True preserva i run colorati (testi parzialmente in rosso)
    wb = openpyxl.load_workbook(SORGENTE, rich_text=True)
    ws = wb.active

    for coord, token in TOKEN_CELLE.items():
        set_cella(ws, coord, token)

    for coord, vecchio, nuovo in SOSTITUZIONI_TESTO:
        sostituisci_testo(ws, coord, vecchio, nuovo)

    percorso_template = DEST_DIR / "template.xlsx"
    wb.save(percorso_template)

    import sys
    sys.path.insert(0, str(RADICE))
    from motore.generatore import ripara_rich_text_salvato
    ripara_rich_text_salvato(percorso_template)

    percorso_config = DEST_DIR / "config.json"
    percorso_config.write_text(
        json.dumps(CONFIG, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("Template creato:", percorso_template)
    print("Config creato:  ", percorso_config)
    print(f"Campi definiti:  {len(CAMPI)}")


if __name__ == "__main__":
    main()
