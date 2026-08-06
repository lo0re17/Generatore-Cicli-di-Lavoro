"""Motore di generazione dei cicli di lavoro.

Approccio a segnaposto (token):
- ogni PN ha un ``template.xlsx`` in cui i campi variabili sono scritti
  come token ``{{NOME}}`` (es. ``{{COMMESSA}}``, ``{{LOTTO_3V000107}}``);
- lo stesso token puo' comparire in piu' celle: viene sostituito ovunque
  (propagazione automatica dei lotti tra distinta base e testo delle fasi);
- un ``config.json`` descrive i campi (etichetta, tipo, obbligatorieta')
  per costruire il form e per sapere come scrivere il valore nella cella.

Il template viene copiato e modificato con openpyxl, che preserva
logo/immagini, stili e impostazioni di stampa (A4 orizzontale, scala, area).
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field as dc_field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as _XLImage
from openpyxl.utils import range_boundaries

from . import ordine_interno as oi

_RE_TOKEN = re.compile(r"\{\{[^{}]+\}\}")
_CARATTERI_VIETATI_FILE = re.compile(r'[\\/:*?"<>|]')

TIPI_TIPIZZATI = {"intero", "data"}  # tipi scritti come valore Excel nativo


# --------------------------------------------------------------------------- #
# Configurazione PN
# --------------------------------------------------------------------------- #
@dataclass
class Campo:
    token: str
    etichetta: str
    tipo: str = "testo"          # testo | intero | data | ordine_interno | lotto
    obbligatorio: bool = False
    default: str = ""
    gruppo: str = ""             # per raggruppare i campi nel form
    aiuto: str = ""              # testo di aiuto opzionale
    fisso: bool = False          # campo con valore fisso dal template, non chiesto nel form


@dataclass(frozen=True)
class NonApplicabile:
    """Segnaposto per un campo escluso volontariamente dall'operatore
    (es. '-' o 'N/A' quando il dato non si applica a questo ciclo).

    Bypassa la validazione di tipo/obbligatorieta' e viene scritto come
    testo letterale nella cella, indipendentemente dal tipo del campo
    (anche su celle 'data' o 'intero')."""
    testo: str = "-"


@dataclass
class InterventoFase:
    """Compilazione di una fase: data di esecuzione e/o firma operatore."""
    cella_data: str | None = None
    data: Any = None                       # datetime | str | None
    cella_firma: str | None = None
    firma_testo: str | None = None
    firma_immagine: str | None = None      # percorso .png


@dataclass
class ConfigPN:
    pn: str
    descrizione: str
    template: str
    campi: list[Campo]
    pattern_nome_file: str = "{ORDINE_INTERNO}_{PN}.xlsx"
    cartella: Path | None = dc_field(default=None)  # cartella del config

    @classmethod
    def da_file(cls, percorso: str | Path) -> "ConfigPN":
        percorso = Path(percorso)
        dati = json.loads(percorso.read_text(encoding="utf-8"))
        campi = [Campo(**c) for c in dati.get("campi", [])]
        return cls(
            pn=dati["pn"],
            descrizione=dati.get("descrizione", ""),
            template=dati["template"],
            campi=campi,
            pattern_nome_file=dati.get(
                "pattern_nome_file", "{ORDINE_INTERNO}_{PN}.xlsx"
            ),
            cartella=percorso.parent,
        )

    def percorso_template(self) -> Path:
        base = self.cartella or Path(".")
        return (base / self.template).resolve()

    def campo_per_tipo(self, tipo: str) -> Campo | None:
        for c in self.campi:
            if c.tipo == tipo:
                return c
        return None

    def mappa_token(self) -> dict[str, Campo]:
        return {c.token: c for c in self.campi}


# --------------------------------------------------------------------------- #
# Conversione dei valori
# --------------------------------------------------------------------------- #
def _parse_data(valore: Any) -> datetime | None:
    """Interpreta una data da vari formati; None se vuota."""
    if valore is None or valore == "":
        return None
    if isinstance(valore, datetime):
        return valore
    if isinstance(valore, date):
        return datetime(valore.year, valore.month, valore.day)
    testo = str(valore).strip()
    if not testo:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
                "%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(testo, fmt)
        except ValueError:
            continue
    raise ValueError(f"Data non riconosciuta: '{valore}' (usa gg/mm/aaaa o gg/mm/aa)")


def _valore_stringa(campo: Campo, valore: Any) -> str:
    """Rappresentazione testuale del valore (per sostituzione in testo)."""
    if isinstance(valore, NonApplicabile):
        return valore.testo
    if valore is None:
        return ""
    if campo.tipo == "data":
        d = _parse_data(valore)
        return d.strftime("%d/%m/%Y") if d else ""
    if campo.tipo == "intero":
        testo = str(valore).strip()
        return testo  # gia' numerico o stringa numerica
    return str(valore).strip()


def _valore_tipizzato(campo: Campo, valore: Any) -> Any:
    """Valore nativo Excel per una cella che contiene solo il token."""
    if isinstance(valore, NonApplicabile):
        return valore.testo
    if campo.tipo == "data":
        return _parse_data(valore)  # datetime o None
    if campo.tipo == "intero":
        testo = str(valore).strip()
        if testo == "":
            return None
        try:
            return int(testo)
        except ValueError:
            return testo  # non numerico: lascio il testo
    return _valore_stringa(campo, valore)


# --------------------------------------------------------------------------- #
# Data e firma nelle celle
# --------------------------------------------------------------------------- #
_EMU_PX = 9525  # EMU per pixel a 96 DPI


def _zona_unita(ws, cella: str) -> tuple[int, int, int, int]:
    """Ritorna (riga_min, col_min, riga_max, col_max) 1-based della cella,
    espandendo all'eventuale zona unita che la contiene."""
    from openpyxl.utils.cell import coordinate_to_tuple

    r0, c0 = coordinate_to_tuple(cella)
    for mr in ws.merged_cells.ranges:
        mc0, mr0, mc1, mr1 = range_boundaries(str(mr))
        if mr0 <= r0 <= mr1 and mc0 <= c0 <= mc1:
            return mr0, mc0, mr1, mc1
    return r0, c0, r0, c0


def scrivi_data_cella(ws, cella: str, data: Any) -> None:
    """Scrive la data come TESTO gg/mm/aaaa (evita '####' nelle celle unite
    strette, dove Excel valuta la larghezza della sola prima colonna)."""
    d = _parse_data(data)
    if d is not None:
        ws[cella] = d.strftime("%d/%m/%Y")


def _px_colonna(ws, indice_col: int) -> int:
    # Le definizioni di colonna possono coprire un intervallo (min..max):
    # openpyxl le indicizza solo sulla prima lettera, quindi si scandisce.
    larghezza = None
    for dim in ws.column_dimensions.values():
        if dim.min and dim.max and dim.min <= indice_col <= dim.max and dim.width:
            larghezza = dim.width
            break
    if larghezza is None:
        larghezza = ws.sheet_format.defaultColWidth or 8.43
    return int(round(larghezza * 7)) + 5


def _px_riga(ws, indice_riga: int) -> int:
    dim = ws.row_dimensions.get(indice_riga)
    altezza = dim.height if dim and dim.height else (ws.sheet_format.defaultRowHeight or 15.0)
    return int(round(altezza * 4 / 3))


def inserisci_immagine(ws, cella: str, percorso: str, inset_px: int = 3) -> None:
    """Inserisce un'immagine nella casella (anche unita) preservando le
    proporzioni, centrata, con dimensione fissa (OneCellAnchor + ext)."""
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D

    r_min, c_min, r_max, c_max = _zona_unita(ws, cella)
    box_w = sum(_px_colonna(ws, c) for c in range(c_min, c_max + 1)) - 2 * inset_px
    box_h = sum(_px_riga(ws, r) for r in range(r_min, r_max + 1)) - 2 * inset_px
    box_w, box_h = max(box_w, 4), max(box_h, 4)

    img = _XLImage(percorso)
    nat_w = img.width or box_w
    nat_h = img.height or box_h
    scala = min(box_w / nat_w, box_h / nat_h)
    disp_w = max(int(nat_w * scala), 1)
    disp_h = max(int(nat_h * scala), 1)

    off_x = (inset_px + (box_w - disp_w) // 2) * _EMU_PX
    off_y = (inset_px + (box_h - disp_h) // 2) * _EMU_PX
    marker = AnchorMarker(col=c_min - 1, row=r_min - 1, colOff=off_x, rowOff=off_y)
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(cx=disp_w * _EMU_PX, cy=disp_h * _EMU_PX),
    )
    ws.add_image(img)


def applica_intervento(ws, intervento: "InterventoFase") -> None:
    if intervento.cella_data and intervento.data:
        scrivi_data_cella(ws, intervento.cella_data, intervento.data)
    if intervento.cella_firma:
        if intervento.firma_immagine:
            inserisci_immagine(ws, intervento.cella_firma, intervento.firma_immagine)
        elif intervento.firma_testo:
            ws[intervento.cella_firma] = intervento.firma_testo


# --------------------------------------------------------------------------- #
# Generazione singola
# --------------------------------------------------------------------------- #
def _sostituisci_in_cella(cella, valore_stringa_cella, mappa: dict[str, Campo],
                          valori: dict[str, Any]) -> None:
    """Applica le sostituzioni dei token presenti in una cella."""
    token_presenti = _RE_TOKEN.findall(valore_stringa_cella)
    if not token_presenti:
        return

    # Caso 1: la cella contiene SOLO un token noto e tipizzato -> valore nativo.
    solo = valore_stringa_cella.strip()
    if (
        len(token_presenti) == 1
        and solo == token_presenti[0]
        and token_presenti[0] in mappa
        and mappa[token_presenti[0]].tipo in TIPI_TIPIZZATI
    ):
        campo = mappa[token_presenti[0]]
        cella.value = _valore_tipizzato(campo, valori.get(campo.token))
        return

    # Caso 2: sostituzione testuale (uno o piu' token, anche dentro una frase).
    nuovo = valore_stringa_cella
    for tok in set(token_presenti):
        campo = mappa.get(tok)
        if campo is None:
            continue  # token sconosciuto: lasciato invariato
        nuovo = nuovo.replace(tok, _valore_stringa(campo, valori.get(campo.token)))
    cella.value = nuovo if nuovo != "" else None


def _sostituisci_in_rich_text(cella, rich, mappa: dict[str, Campo],
                              valori: dict[str, Any]) -> None:
    """Sostituisce i token dentro un valore rich text (testo con formattazioni
    parziali, es. parole in rosso) preservando i run e i loro colori.

    I token vengono sostituiti run per run: i template creati dal tool
    scrivono ogni token per intero dentro un singolo run.
    """
    from openpyxl.cell.rich_text import TextBlock

    for i, parte in enumerate(rich):
        testo = parte.text if isinstance(parte, TextBlock) else str(parte)
        if "{{" not in testo:
            continue
        for tok in set(_RE_TOKEN.findall(testo)):
            campo = mappa.get(tok)
            if campo is None:
                continue
            testo = testo.replace(tok, _valore_stringa(campo, valori.get(campo.token)))
        if isinstance(parte, TextBlock):
            parte.text = testo
        else:
            rich[i] = testo
    cella.value = rich


def ripara_rich_text_salvato(percorso: str | Path) -> None:
    """Post-processa un .xlsx salvato da openpyxl in modalita' rich text.

    openpyxl 3.1.x non aggiunge ``xml:space="preserve"`` ai run inline:
    Excel rifiuta di aprire i file quando un run contiene un a-capo.
    Qui si marca preserve su tutti i ``<t>`` (innocuo dove non serve).
    """
    import zipfile

    percorso = Path(percorso)
    entrate: dict[str, bytes] = {}
    with zipfile.ZipFile(percorso) as z:
        for nome in z.namelist():
            dati = z.read(nome)
            if nome.startswith("xl/worksheets/") and nome.endswith(".xml") \
                    or nome == "xl/sharedStrings.xml":
                dati = dati.replace(b"<t>", b'<t xml:space="preserve">')
            entrate[nome] = dati
    tmp = percorso.with_suffix(".tmp_riparazione")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for nome, dati in entrate.items():
            z.writestr(nome, dati)
    tmp.replace(percorso)


def genera_ciclo(config: ConfigPN, valori: dict[str, Any], percorso_out: str | Path,
                 interventi: list["InterventoFase"] | None = None) -> Path:
    """Genera un singolo ciclo .xlsx a partire dal template e dai valori.

    ``valori`` e' un dizionario token -> valore (es. {"{{COMMESSA}}": "..."}).
    ``interventi`` (opzionale) compila data esecuzione e/o firma nelle fasi.

    Il workbook e' caricato in modalita' rich text: i testi con formattazioni
    parziali (es. valori in rosso dentro una frase) restano intatti.
    """
    from openpyxl.cell.rich_text import CellRichText

    mappa = config.mappa_token()
    wb = openpyxl.load_workbook(config.percorso_template(), rich_text=True)
    for ws in wb.worksheets:
        for riga in ws.iter_rows():
            for cella in riga:
                v = cella.value
                if isinstance(v, CellRichText):
                    if "{{" in str(v):
                        _sostituisci_in_rich_text(cella, v, mappa, valori)
                elif isinstance(v, str) and "{{" in v:
                    _sostituisci_in_cella(cella, v, mappa, valori)
    if interventi:
        ws0 = wb.worksheets[0]
        for intervento in interventi:
            applica_intervento(ws0, intervento)
    percorso_out = Path(percorso_out)
    percorso_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(percorso_out)
    ripara_rich_text_salvato(percorso_out)
    return percorso_out


# --------------------------------------------------------------------------- #
# Generazione batch
# --------------------------------------------------------------------------- #
def _sanitizza_nome(nome: str) -> str:
    return _CARATTERI_VIETATI_FILE.sub("-", nome).strip()


def _nome_file(config: ConfigPN, valori: dict[str, Any], suffisso: str = "") -> str:
    """Rende il pattern del nome file usando PN + valori dei token.

    ``suffisso`` (es. uno shipset "SS1") viene aggiunto prima dell'estensione."""
    contesto = {"PN": config.pn}
    for c in config.campi:
        nome_chiave = c.token.strip("{}")  # {{ORDINE_INTERNO}} -> ORDINE_INTERNO
        contesto[nome_chiave] = _valore_stringa(c, valori.get(c.token))

    def sostituisci(m: re.Match) -> str:
        chiave = m.group(1)
        return str(contesto.get(chiave, m.group(0)))

    nome = re.sub(r"\{([A-Za-z0-9_]+)\}", sostituisci, config.pattern_nome_file)
    suffisso = suffisso.strip()
    if suffisso:
        p = Path(nome)
        nome = f"{p.stem}_{suffisso}{p.suffix}"
    return _sanitizza_nome(nome)


def genera_batch(
    config: ConfigPN,
    valori_comuni: dict[str, Any],
    quantita: int,
    cartella_out: str | Path,
    pezzi_override: dict[int, Any] | None = None,
    interventi: list["InterventoFase"] | None = None,
    suffisso_nome: str = "",
) -> list[Path]:
    """Genera ``quantita`` cicli incrementando l'Ordine Interno.

    - ``valori_comuni``: valori condivisi da tutti i cicli (token -> valore);
      deve contenere il primo Ordine Interno.
    - ``pezzi_override``: {indice_ciclo (0-based): nr_pezzi} per dare all'ultimo
      ciclo di un lotto un numero di pezzi diverso.
    - ``interventi``: data/firma da applicare alle fasi (uguali per tutti i cicli).
    - ``suffisso_nome``: testo libero (es. "SS1") aggiunto al nome file.
    """
    if quantita < 1:
        return []

    campo_oi = config.campo_per_tipo("ordine_interno")
    if campo_oi is None:
        raise ValueError("Config senza campo di tipo 'ordine_interno'.")
    campo_pezzi = config.campo_per_tipo("intero")  # convenzione: Nr. Pezzi

    base_oi = str(valori_comuni.get(campo_oi.token, "")).strip()
    if not base_oi:
        raise ValueError("Ordine Interno di partenza mancante.")
    serie_oi = oi.serie(base_oi, quantita)

    cartella_out = Path(cartella_out)
    prodotti: list[Path] = []
    pezzi_override = pezzi_override or {}

    for i in range(quantita):
        valori = dict(valori_comuni)
        valori[campo_oi.token] = serie_oi[i]
        if campo_pezzi is not None and i in pezzi_override:
            valori[campo_pezzi.token] = pezzi_override[i]
        nome = _nome_file(config, valori, suffisso_nome)
        prodotti.append(
            genera_ciclo(config, valori, cartella_out / nome, interventi=interventi)
        )

    return prodotti


# --------------------------------------------------------------------------- #
# Validazione input
# --------------------------------------------------------------------------- #
def valida(config: ConfigPN, valori: dict[str, Any]) -> list[str]:
    """Ritorna la lista degli errori (vuota se tutto ok)."""
    errori: list[str] = []
    for c in config.campi:
        if c.fisso:
            continue  # valore dal template, non validato
        grezzo = valori.get(c.token, "")
        if isinstance(grezzo, NonApplicabile):
            continue  # escluso volontariamente: nessun controllo di tipo
        vuoto = grezzo is None or str(grezzo).strip() == ""
        if c.obbligatorio and vuoto:
            errori.append(f"Campo obbligatorio mancante: {c.etichetta}")
            continue
        if vuoto:
            continue
        if c.tipo == "data":
            try:
                _parse_data(grezzo)
            except ValueError as e:
                errori.append(f"{c.etichetta}: {e}")
        elif c.tipo == "intero":
            try:
                int(str(grezzo).strip())
            except ValueError:
                errori.append(f"{c.etichetta}: deve essere un numero intero")
    return errori
