"""Rilevamento delle fasi di lavoro in un ciclo.

Individua la tabella delle fasi (intestazione con "Fase", "Data esecuzione",
"Firma operatore") e, per ogni fase, la cella dove scrivere la data di
esecuzione e la firma dell'operatore. Usato sia per compilare date/firme sia
dal futuro creatore di template.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.utils import range_boundaries


@dataclass
class Fase:
    riga: int
    numero: str
    reparto: str
    descrizione: str
    cella_data: str
    cella_firma: str

    def etichetta(self) -> str:
        desc = (self.descrizione or "").strip().replace("\n", " ")
        if len(desc) > 60:
            desc = desc[:57] + "…"
        base = f"Fase {self.numero}"
        if self.reparto:
            base += f" [{self.reparto}]"
        return f"{base} — {desc}" if desc else base


def _anchor(ws, r: int, c: int) -> str:
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(mr))
        if min_r <= r <= max_r and min_c <= c <= max_c:
            return ws.cell(row=min_r, column=min_c).coordinate
    return ws.cell(row=r, column=c).coordinate


def _trova_intestazione(ws) -> tuple[int, int, int, int] | None:
    """Ritorna (riga_header, col_fase, col_data, col_firma) o None."""
    for riga in ws.iter_rows(min_row=1, max_row=ws.max_row):
        col_fase = col_data = col_firma = None
        for cella in riga:
            v = cella.value
            if not isinstance(v, str):
                continue
            testo = v.strip().lower()
            if testo == "fase":
                col_fase = cella.column
            elif testo.startswith("data") and "esecuzione" in testo:
                col_data = cella.column
            elif testo.startswith("firma"):
                col_firma = cella.column
        if col_fase and col_data and col_firma:
            return riga[0].row, col_fase, col_data, col_firma
    return None


def rileva_fasi(ws) -> list[Fase]:
    intest = _trova_intestazione(ws)
    if intest is None:
        return []
    riga_header, col_fase, col_data, col_firma = intest

    # colonna descrizione: cella "Descrizione" nella riga di intestazione
    col_desc = None
    for cella in ws[riga_header]:
        if isinstance(cella.value, str) and cella.value.strip().lower() == "descrizione":
            col_desc = cella.column
            break

    fasi: list[Fase] = []
    ultimo_numero: int | None = None
    for r in range(riga_header + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=col_fase).value
        if val is None or str(val).strip() == "":
            continue
        # numero fase: se formula tipo =A53+10, calcola dal precedente
        testo = str(val).strip()
        if testo.startswith("="):
            numero = str((ultimo_numero + 10)) if ultimo_numero is not None else testo
        else:
            numero = testo
        try:
            ultimo_numero = int(float(numero))
        except ValueError:
            pass

        reparto = ws.cell(row=r, column=col_fase + 1).value
        reparto = str(reparto).strip() if reparto is not None else ""
        descr = ""
        if col_desc:
            d = ws.cell(row=r, column=col_desc).value
            descr = str(d).strip() if d is not None else ""

        fasi.append(Fase(
            riga=r,
            numero=numero,
            reparto=reparto,
            descrizione=descr,
            cella_data=_anchor(ws, r, col_data),
            cella_firma=_anchor(ws, r, col_firma),
        ))
    return fasi
