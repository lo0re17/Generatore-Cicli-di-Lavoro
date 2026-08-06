"""Riconoscimento automatico della struttura di un RIQ compilato.

Analizza un Rapporto Ispezione Qualita' esistente (famiglia "Mod 02-01_riq_CQ")
e propone: i campi di testata (token), la tabella misure (righe/colonne) e
il blocco esito (data, stato, firma CQ). Segue lo stesso schema di
``rileva_campi.py`` per l'ODL, adattato alla struttura del RIQ.

Uso:
    proposta = analizza_riq("riq_compilato.xlsx")
    # ... eventuale revisione (GUI wizard) ...
    costruisci_template_riq(proposta, "templates/")
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple


# --------------------------------------------------------------------------- #
# Modello della proposta
# --------------------------------------------------------------------------- #
@dataclass
class CampoRIQProposto:
    token: str
    etichetta: str
    tipo: str
    cella: str = ""
    valore_attuale: str = ""
    gruppo: str = "Testata"
    aiuto: str = ""
    obbligatorio: bool = False
    default: str = ""
    incluso: bool = True
    fisso: bool = False


@dataclass
class PropostaRIQ:
    sorgente: str
    pn: str
    descrizione: str
    campi: list[CampoRIQProposto] = field(default_factory=list)
    riga_inizio_misure: int = 0
    riga_fine_misure: int = 0
    colonne_misure: dict[str, str] = field(default_factory=dict)
    righe_misura_preview: list[dict] = field(default_factory=list)
    riga_esito: int = 0
    colonna_data_esito: str = ""
    colonna_stato_esito: str = ""
    cella_firma_cq: str = ""
    riga_non_conformita: int = 0
    cella_rnc: str = ""
    cella_concessione: str = ""
    cella_redatto: str = ""
    avvisi: list[str] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Utilita' foglio (identiche a rileva_campi.py)
# --------------------------------------------------------------------------- #
def _mappa_merge(ws) -> list[tuple[int, int, int, int]]:
    return [range_boundaries(str(mr)) for mr in ws.merged_cells.ranges]


def _anchor(ws, r: int, c: int, merges) -> tuple[int, int]:
    for mc0, mr0, mc1, mr1 in merges:
        if mr0 <= r <= mr1 and mc0 <= c <= mc1:
            return mr0, mc0
    return r, c


def _fine_merge_dx(ws, r: int, c: int, merges) -> int:
    for mc0, mr0, mc1, mr1 in merges:
        if mr0 <= r <= mr1 and mc0 <= c <= mc1:
            return mc1
    return c


def _valore_dx(ws, r: int, c: int, merges, max_scan: int = 8):
    c_fine = _fine_merge_dx(ws, r, c, merges)
    for cc in range(c_fine + 1, min(c_fine + 1 + max_scan, ws.max_column + 1)):
        ar, ac = _anchor(ws, r, cc, merges)
        cella = ws.cell(row=ar, column=ac)
        return cella.coordinate, cella.value
    return None, None


def _cerca_etichetta(ws, testo: str, esatta: bool = False) -> list[tuple[int, int]]:
    testo_low = testo.strip().lower()
    trovate = []
    for riga in ws.iter_rows():
        for cella in riga:
            v = cella.value
            if not isinstance(v, str):
                continue
            v_low = v.strip().lower().replace("\n", " ")
            if (esatta and v_low == testo_low) or (not esatta and testo_low in v_low):
                trovate.append((cella.row, cella.column))
    return trovate


def _fmt(valore) -> str:
    return "" if valore is None else str(valore)


def _slug(testo: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(testo).strip()).strip("_").upper()
    return s or "CAMPO"


# --------------------------------------------------------------------------- #
# Analisi
# --------------------------------------------------------------------------- #
def analizza_riq(percorso: str | Path) -> PropostaRIQ:
    percorso = Path(percorso)
    wb = openpyxl.load_workbook(percorso)
    ws = wb.active
    merges = _mappa_merge(ws)
    prop = PropostaRIQ(sorgente=str(percorso), pn="", descrizione="")

    def aggiungi(campo: CampoRIQProposto) -> None:
        prop.campi.append(campo)

    def campo_testata(etich: str, token: str, tipo: str, obbl: bool = False,
                      aiuto: str = "", esatta: bool = False) -> str | None:
        posizioni = _cerca_etichetta(ws, etich, esatta=esatta)
        if not posizioni:
            prop.avvisi.append(f"Etichetta '{etich}' non trovata.")
            return None
        r, c = posizioni[0]
        coord, v = _valore_dx(ws, r, c, merges)
        if coord is None:
            prop.avvisi.append(f"Nessuna cella valore accanto a '{etich}'.")
            return None
        aggiungi(CampoRIQProposto(
            token=token, etichetta=etich, tipo=tipo, cella=coord,
            valore_attuale=_fmt(v), obbligatorio=obbl, aiuto=aiuto))
        return coord

    campo_testata("RIQ:", "{{RIQ}}", "testo")
    campo_testata("Protocollo Nr:", "{{PROTOCOLLO}}", "testo")
    campo_testata("Cliente:", "{{CLIENTE}}", "testo")
    campo_testata("Commessa:", "{{COMMESSA}}", "testo", obbl=True)
    campo_testata("Ordine acquisto cliente", "{{ORDINE_ACQUISTO}}", "testo")
    campo_testata("Descrizione:", "{{DESCRIZIONE}}", "testo")
    campo_testata("Ordine di lavoro n", "{{ORDINE_INTERNO}}", "ordine_interno",
                  obbl=True, aiuto="Deve coincidere con l'Ordine Interno dell'ODL.")
    campo_testata("Lotto n", "{{LOTTO}}", "testo")

    pn_pos = _cerca_etichetta(ws, "PN:", esatta=False)
    if pn_pos:
        r, c = pn_pos[0]
        coord, v = _valore_dx(ws, r, c, merges)
        prop.pn = _fmt(v).strip()
        if coord:
            aggiungi(CampoRIQProposto("{{PN}}", "PN", "testo", cella=coord,
                                      valore_attuale=_fmt(v), obbligatorio=True))
    else:
        prop.avvisi.append("Etichetta 'PN:' non trovata: PN da inserire a mano.")

    campo_testata("DWG:", "{{DWG}}", "testo")
    campo_testata("Quantit", "{{QUANTITA_PRODOTTE}}", "intero")
    campo_testata("% campionamento", "{{PERC_CAMPIONAMENTO}}", "testo")

    # Descrizione RIQ (usata come descrizione della cartella template)
    desc_campo = next((c for c in prop.campi if c.token == "{{DESCRIZIONE}}"), None)
    if desc_campo:
        prop.descrizione = desc_campo.valore_attuale

    # ---------------- Tabella "ISPEZIONI DA ESEGUIRE"
    header_labels = {
        "posizione": "posizione caratteristica",
        "caratteristica": "caratteristica",
        "modalita_verifica": "modalit",
        "strumento": "strumento utilizzato",
        "accettato": "accettato",
        "scartato": "scartato",
        "risultato": "risultato rilevato",
    }
    header_row = None
    trovate_col: dict[str, int] = {}
    for riga in ws.iter_rows():
        for cella in riga:
            v = cella.value
            if not isinstance(v, str):
                continue
            v_low = v.strip().lower().replace("\n", " ")
            for chiave, etich in header_labels.items():
                if etich in v_low:
                    trovate_col[chiave] = cella.column
                    header_row = cella.row
        if len(trovate_col) >= 5:
            break

    if header_row and "posizione" in trovate_col:
        from openpyxl.utils import get_column_letter
        prop.colonne_misure = {k: get_column_letter(v) for k, v in trovate_col.items()}
        col_pos = trovate_col["posizione"]
        r = header_row + 1
        prop.riga_inizio_misure = r
        ultima_valida = r - 1
        while r <= ws.max_row:
            ar, ac = _anchor(ws, r, col_pos, merges)
            val = ws.cell(row=ar, column=ac).value
            if val is None or str(val).strip() == "":
                break
            ultima_valida = r
            riga_preview = {}
            for chiave, col in trovate_col.items():
                riga_preview[chiave] = _fmt(ws.cell(row=r, column=col).value)
            riga_preview["riga"] = r
            prop.righe_misura_preview.append(riga_preview)
            r += 1
        prop.riga_fine_misure = ultima_valida
    else:
        prop.avvisi.append("Tabella misure non rilevata (intestazione non trovata).")

    # ---------------- Blocco esito: cerca "Controllo Qualit" nella stessa riga di uno stato
    stati_possibili = {"conforme", "non conforme"}
    riga_esito = None
    col_stato = None
    for riga in ws.iter_rows():
        for cella in riga:
            v = cella.value
            if isinstance(v, str) and v.strip().lower() in stati_possibili:
                riga_esito = cella.row
                col_stato = cella.column
                break
        if riga_esito:
            break

    if riga_esito:
        from openpyxl.utils import get_column_letter
        prop.riga_esito = riga_esito
        prop.colonna_stato_esito = get_column_letter(col_stato)
        # data: prima cella con un datetime o formato data sulla stessa riga
        for cella in ws[riga_esito]:
            if cella.value is not None and cella.column < col_stato:
                prop.colonna_data_esito = get_column_letter(cella.column)
                break
        # firma CQ: cerca "Controllo Qualit" sulla stessa riga, la firma va
        # nella zona a destra (immagine ancorata li' o cella accanto)
        for cc in _cerca_etichetta(ws, "Controllo Qualit"):
            if cc[0] == riga_esito:
                r, c = cc
                c_fine = _fine_merge_dx(ws, r, c, merges)
                ar, ac = _anchor(ws, r, c_fine + 1, merges)
                prop.cella_firma_cq = ws.cell(row=ar, column=ac).coordinate
                break
    else:
        prop.avvisi.append("Riga esito (CONFORME/NON CONFORME) non trovata.")

    # ---------------- RNC / Richiesta Concessione
    rnc_pos = _cerca_etichetta(ws, "N. RNC")
    conc_pos = _cerca_etichetta(ws, "Richiesta Concessione")
    if rnc_pos:
        r, c = rnc_pos[0]
        prop.riga_non_conformita = r
        coord, _ = _valore_dx(ws, r, c, merges, max_scan=2)
        if coord:
            prop.cella_rnc = coord
    if conc_pos:
        r, c = conc_pos[0]
        coord, _ = _valore_dx(ws, r, c, merges, max_scan=2)
        if coord:
            prop.cella_concessione = coord

    # ---------------- Redatto da
    redatto_pos = _cerca_etichetta(ws, "Redatto da")
    if redatto_pos:
        r, c = redatto_pos[0]
        ar, ac = _anchor(ws, r, c, merges)
        prop.cella_redatto = ws.cell(row=ar, column=ac).coordinate

    return prop


# --------------------------------------------------------------------------- #
# Costruzione del template dalla proposta
# --------------------------------------------------------------------------- #
def costruisci_template_riq(prop: PropostaRIQ, cartella_templates: str | Path,
                            pattern_nome_file: str | None = None) -> Path:
    """Applica la proposta: scrive template_riq.xlsx + config_riq.json nella
    cartella del PN (stessa cartella usata dal template ODL)."""
    if not prop.pn:
        raise ValueError("PN mancante nella proposta.")
    dest = Path(cartella_templates) / prop.pn
    dest.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(prop.sorgente, rich_text=True)
    ws = wb.active
    merges = _mappa_merge(ws)

    campi_config = []
    for campo in prop.campi:
        if not campo.incluso:
            continue
        if campo.cella:
            r, c = coordinate_to_tuple(campo.cella)
            ar, ac = _anchor(ws, r, c, merges)
            ws.cell(row=ar, column=ac).value = campo.token
        campi_config.append({
            "token": campo.token,
            "etichetta": campo.etichetta,
            "tipo": campo.tipo,
            "obbligatorio": campo.obbligatorio,
            "default": campo.default,
            "gruppo": campo.gruppo,
            "aiuto": campo.aiuto,
            "fisso": campo.fisso,
        })

    nome_template = "template_riq.xlsx"
    wb.save(dest / nome_template)
    from motore.generatore import ripara_rich_text_salvato
    ripara_rich_text_salvato(dest / nome_template)

    config = {
        "pn": prop.pn,
        "descrizione": prop.descrizione,
        "template": nome_template,
        "pattern_nome_file": pattern_nome_file
            or "{ORDINE_INTERNO}_RIQ_OR_{PN}.xlsx",
        "campi": campi_config,
        "riga_inizio_misure": prop.riga_inizio_misure,
        "riga_fine_misure": prop.riga_fine_misure,
        "colonne_misure": {
            "posizione": prop.colonne_misure.get("posizione", "A"),
            "caratteristica": prop.colonne_misure.get("caratteristica", "B"),
            "modalita_verifica": prop.colonne_misure.get("modalita_verifica", "F"),
            "strumento": prop.colonne_misure.get("strumento", "G"),
            "accettato": prop.colonne_misure.get("accettato", "H"),
            "scartato": prop.colonne_misure.get("scartato", "I"),
            "risultato": prop.colonne_misure.get("risultato", "J"),
        },
        "blocco_esito": {
            "riga": prop.riga_esito,
            "colonna_data": prop.colonna_data_esito or "A",
            "colonna_stato": prop.colonna_stato_esito or "D",
            "cella_firma_cq": prop.cella_firma_cq or f"H{prop.riga_esito}",
        },
        "blocco_non_conformita": {
            "riga": prop.riga_non_conformita,
            "cella_rnc": prop.cella_rnc,
            "cella_concessione": prop.cella_concessione,
        } if prop.riga_non_conformita else None,
        "blocco_redatto_approvato": {
            "cella_redatto": prop.cella_redatto,
            "cella_approvato": "",
        } if prop.cella_redatto else None,
    }
    (dest / "config_riq.json").write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return dest
