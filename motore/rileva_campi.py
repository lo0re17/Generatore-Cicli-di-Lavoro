"""Riconoscimento automatico dei campi variabili in un ciclo compilato.

Analizza un ciclo di lavoro esistente (famiglia di layout 01-03-cl-RT) e
propone la mappatura celle -> token per crearne il template. Le euristiche
si basano sulle etichette stabili del modulo ("Commessa", "Ordine Interno",
"Lotto", ...) e sono robuste rispetto a piccole differenze di posizione.

Uso:
    proposta = analizza("ciclo_compilato.xlsx")
    # ... eventuale revisione della proposta (GUI wizard) ...
    costruisci_template(proposta, "templates/")
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries


# --------------------------------------------------------------------------- #
# Modello della proposta
# --------------------------------------------------------------------------- #
@dataclass
class CampoProposto:
    token: str
    etichetta: str
    tipo: str                    # testo | intero | data | ordine_interno | lotto
    cella: str = ""              # cella da tokenizzare (vuota se solo sostituzioni)
    valore_attuale: str = ""
    gruppo: str = ""
    aiuto: str = ""
    obbligatorio: bool = False
    default: str = ""
    # sostituzioni testuali extra: [(cella, testo_da_sostituire), ...]
    sostituzioni: list[tuple[str, str]] = field(default_factory=list)
    incluso: bool = True         # il wizard permette di escluderlo
    fisso: bool = False          # valore fisso dal template, non richiesto nel form


@dataclass
class Proposta:
    sorgente: str
    pn: str
    descrizione: str
    campi: list[CampoProposto] = field(default_factory=list)
    avvisi: list[str] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Utilita' foglio
# --------------------------------------------------------------------------- #
def _mappa_merge(ws) -> list[tuple[int, int, int, int]]:
    return [range_boundaries(str(mr)) for mr in ws.merged_cells.ranges]


def _anchor(ws, r: int, c: int, merges) -> tuple[int, int]:
    for mc0, mr0, mc1, mr1 in merges:
        if mr0 <= r <= mr1 and mc0 <= c <= mc1:
            return mr0, mc0
    return r, c


def _fine_merge_dx(ws, r: int, c: int, merges) -> int:
    """Colonna finale della zona unita che contiene (r, c)."""
    for mc0, mr0, mc1, mr1 in merges:
        if mr0 <= r <= mr1 and mc0 <= c <= mc1:
            return mc1
    return c


def _valore_dx(ws, r: int, c: int, merges, max_scan: int = 12):
    """Prima cella (anchor) a destra della zona unita di (r,c). Ritorna
    (coordinate, valore) — la cella puo' essere vuota."""
    c_fine = _fine_merge_dx(ws, r, c, merges)
    for cc in range(c_fine + 1, min(c_fine + 1 + max_scan, ws.max_column + 1)):
        ar, ac = _anchor(ws, r, cc, merges)
        cella = ws.cell(row=ar, column=ac)
        return cella.coordinate, cella.value
    return None, None


def _cerca_etichetta(ws, testo: str, esatta: bool = True) -> list[tuple[int, int]]:
    """Celle il cui valore corrisponde all'etichetta (case-insensitive)."""
    testo_low = testo.strip().lower()
    trovate = []
    for riga in ws.iter_rows():
        for cella in riga:
            v = cella.value
            if not isinstance(v, str):
                continue
            v_low = v.strip().lower()
            if (esatta and v_low == testo_low) or (not esatta and testo_low in v_low):
                trovate.append((cella.row, cella.column))
    return trovate


def _fmt(valore) -> str:
    if valore is None:
        return ""
    return str(valore)


_RE_ORDINE_INTERNO = re.compile(r"^\s*\d+\s*-\s*\d+\s*$")


def _slug(testo: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(testo).strip()).strip("_").upper()
    return s or "CAMPO"


# --------------------------------------------------------------------------- #
# Analisi
# --------------------------------------------------------------------------- #
def analizza(percorso: str | Path) -> Proposta:
    percorso = Path(percorso)
    wb = openpyxl.load_workbook(percorso)
    ws = wb.active
    merges = _mappa_merge(ws)
    prop = Proposta(sorgente=str(percorso), pn="", descrizione="")

    def aggiungi(campo: CampoProposto) -> None:
        prop.campi.append(campo)

    # ---------------- P/N e descrizione (riga "P/N")
    pn_pos = _cerca_etichetta(ws, "P/N")
    if pn_pos:
        r, c = pn_pos[0]
        _, v = _valore_dx(ws, r, c, merges)
        prop.pn = _fmt(v).strip()
        for rr, cc in _cerca_etichetta(ws, "DESCRIZIONE"):
            if rr == r:
                _, dv = _valore_dx(ws, rr, cc, merges)
                prop.descrizione = _fmt(dv).strip()
                break
    else:
        prop.avvisi.append("Etichetta 'P/N' non trovata: PN da inserire a mano.")

    # ---------------- Testata
    def campo_testata(etich: str, token: str, tipo: str, obbl: bool = False,
                      aiuto: str = "") -> str | None:
        """Cerca etichetta e propone il campo. Ritorna la cella valore."""
        posizioni = _cerca_etichetta(ws, etich)
        if not posizioni:
            prop.avvisi.append(f"Etichetta '{etich}' non trovata.")
            return None
        r, c = posizioni[0]
        coord, v = _valore_dx(ws, r, c, merges)
        if coord is None:
            prop.avvisi.append(f"Nessuna cella valore accanto a '{etich}'.")
            return None
        aggiungi(CampoProposto(
            token=token, etichetta=etich, tipo=tipo, cella=coord,
            valore_attuale=_fmt(v), gruppo="Testata", obbligatorio=obbl,
            aiuto=aiuto))
        return coord

    campo_testata("Commessa", "{{COMMESSA}}", "testo", obbl=True)
    cella_oda = campo_testata("Ordine d'Acquisto", "{{ORDINE_ACQUISTO}}", "testo")
    cella_oi = campo_testata(
        "Ordine Interno", "{{ORDINE_INTERNO}}", "ordine_interno", obbl=True,
        aiuto="Formato NNNN-YY. Si incrementa da solo per i cicli successivi.")
    campo_testata("Nr. Pezzi", "{{NR_PEZZI}}", "intero")

    # "Data" generiche: associate a ODA/OI dalla colonna piu' vicina
    date_pos = _cerca_etichetta(ws, "Data")
    if date_pos and (cella_oda or cella_oi):
        def col_di(coord: str | None) -> int | None:
            if not coord:
                return None
            from openpyxl.utils.cell import coordinate_to_tuple
            return coordinate_to_tuple(coord)[1]

        col_oda, col_oi = col_di(cella_oda), col_di(cella_oi)
        for r, c in date_pos:
            coord, v = _valore_dx(ws, r, c, merges)
            if coord is None:
                continue
            from openpyxl.utils.cell import coordinate_to_tuple
            col_val = coordinate_to_tuple(coord)[1]
            if col_oda is not None and abs(col_val - col_oda) <= 2:
                aggiungi(CampoProposto("{{DATA_ODA}}", "Data ODA", "data",
                                       cella=coord, valore_attuale=_fmt(v),
                                       gruppo="Testata", aiuto="gg/mm/aaaa"))
            elif col_oi is not None and abs(col_val - col_oi) <= 2:
                aggiungi(CampoProposto("{{DATA_OI}}", "Data OI", "data",
                                       cella=coord, valore_attuale=_fmt(v),
                                       gruppo="Testata", aiuto="gg/mm/aaaa"))

    campo_testata("Data Consegna", "{{DATA_CONSEGNA}}", "data",
                  aiuto="gg/mm/aaaa (opzionale)")

    # Rev. sulla riga del P/N (evita la Rev. dello Stato Revisioni)
    if pn_pos:
        r_pn = pn_pos[0][0]
        for r, c in _cerca_etichetta(ws, "Rev."):
            if r == r_pn:
                coord, v = _valore_dx(ws, r, c, merges)
                if coord:
                    aggiungi(CampoProposto("{{REV}}", "Rev.", "testo",
                                           cella=coord, valore_attuale=_fmt(v),
                                           gruppo="Testata", default=_fmt(v)))
                break

    # ---------------- Distinta base: colonne "Lotto"
    intestazioni_lotto = _cerca_etichetta(ws, "Lotto")
    righe_gia_usate: set[int] = set()
    for r_int, c_int in intestazioni_lotto:
        # individua le colonne Cod. e Descrizione della stessa tabella
        col_cod = col_desc = None
        for rr, cc in _cerca_etichetta(ws, "Cod."):
            if rr == r_int:
                col_cod = cc
        for rr, cc in _cerca_etichetta(ws, "Descrizione"):
            if rr == r_int:
                col_desc = cc
        if col_cod is None:
            continue
        # scorri le righe sotto l'intestazione finche' c'e' un codice
        for r in range(r_int + 1, ws.max_row + 1):
            if r in righe_gia_usate:
                continue
            ar, ac = _anchor(ws, r, col_cod, merges)
            cod = ws.cell(row=ar, column=ac).value
            if cod is None or str(cod).strip() == "":
                break
            righe_gia_usate.add(r)
            cod_s = str(cod).strip()
            # radice del codice senza suffisso -0000
            radice = re.sub(r"-0+$", "", cod_s)
            token = "{{LOTTO_" + _slug(radice) + "}}"
            ar2, ac2 = _anchor(ws, r, c_int, merges)
            cella_lotto = ws.cell(row=ar2, column=ac2)
            desc = ""
            if col_desc is not None:
                ar3, ac3 = _anchor(ws, r, col_desc, merges)
                desc = _fmt(ws.cell(row=ar3, column=ac3).value).strip()
            etich = f"{radice} — {desc}" if desc else radice
            if len(etich) > 55:
                etich = etich[:52] + "…"
            aggiungi(CampoProposto(
                token=token, etichetta=etich, tipo="lotto",
                cella=cella_lotto.coordinate,
                valore_attuale=_fmt(cella_lotto.value),
                gruppo="Lotti distinta base"))

    # ---------------- Propagazione lotti + "Lotto:"/"Scadenza:" nelle fasi
    celle_lotto_note = {c.cella for c in prop.campi if c.tipo == "lotto"}
    valori_lotto = {c.valore_attuale: c for c in prop.campi
                    if c.tipo == "lotto" and c.valore_attuale}

    # Case-insensitive, tollera abbreviazioni ("Lotto n.:", "Scad.:") e
    # testo prima dell'etichetta ("Vedi Lotto: X"), non solo a inizio cella.
    _RE_LOTTO_LABEL = re.compile(r"(?i)lotto\s*n?\.?\s*:\s*(.*)")
    _RE_SCADENZA_LABEL = re.compile(r"(?i)scad(?:enza)?\.?\s*:\s*(.*)")
    _RE_STOP_VALORE = re.compile(r"(?i)\s{2,}|\bscadenza\b|\bscad\.|\blotto\b")

    def _estrai_valore(coda: str) -> str:
        coda = coda.strip()
        m = _RE_STOP_VALORE.search(coda)
        return (coda[:m.start()] if m else coda).strip(" \t-:")

    n_fase_anon = 0
    for riga in ws.iter_rows():
        for cella in riga:
            v = cella.value
            if not isinstance(v, str) or cella.coordinate in celle_lotto_note:
                continue

            # 1) propagazione: il testo contiene il valore di un lotto noto
            for val, campo in valori_lotto.items():
                if val and val in v and cella.coordinate != campo.cella:
                    # evita falsi positivi su valori molto corti
                    if len(val) >= 3 or f"Lotto: {val}" in v:
                        campo.sostituzioni.append((cella.coordinate, val))

            # righe multiple nella stessa cella (testo con "a capo")
            for riga_testo in v.splitlines() or [v]:
                # 2) "Lotto:" con valore non censito in distinta
                m = _RE_LOTTO_LABEL.search(riga_testo)
                if m:
                    valore = _estrai_valore(m.group(1))
                    if not valore:
                        coord_dx, val_dx = _valore_dx(ws, cella.row, cella.column, merges,
                                                       max_scan=3)
                        valore = _fmt(val_dx).strip() if val_dx else ""
                    if valore not in valori_lotto:
                        n_fase_anon += 1
                        token = "{{LOTTO_FASE_" + str(n_fase_anon) + "}}"
                        aggiungi(CampoProposto(
                            token=token,
                            etichetta=f"Lotto in fase (cella {cella.coordinate})",
                            tipo="lotto", valore_attuale=valore,
                            gruppo="Lotti nelle fasi",
                            sostituzioni=[(cella.coordinate,
                                          valore if valore else "__APPEND__")]))

                # 3) "Scadenza:" con o senza valore
                m = _RE_SCADENZA_LABEL.search(riga_testo)
                if m:
                    valore_scad = _estrai_valore(m.group(1))
                    token = "{{SCADENZA_" + str(len(prop.campi)) + "}}"
                    if valore_scad:
                        sost = [(cella.coordinate, valore_scad)]
                    else:
                        sost = [(cella.coordinate, "__APPEND__")]
                    aggiungi(CampoProposto(
                        token=token,
                        etichetta=f"Scadenza (cella {cella.coordinate})",
                        tipo="data", valore_attuale=valore_scad,
                        gruppo="Lotti nelle fasi", aiuto="gg/mm/aaaa (opzionale)",
                        sostituzioni=sost))

    return prop


# --------------------------------------------------------------------------- #
# Costruzione del template dalla proposta
# --------------------------------------------------------------------------- #
def _sostituisci_testo_cella(cella, testo: str, token: str) -> None:
    """Sostituisce ``testo`` con ``token`` nella cella, gestendo sia stringhe
    semplici sia rich text (run per run, colori preservati). Il valore
    speciale __APPEND__ accoda il token in fondo al testo esistente."""
    from openpyxl.cell.rich_text import CellRichText, TextBlock

    v = cella.value
    if isinstance(v, CellRichText):
        if testo == "__APPEND__":
            v.append(" " + token)
            cella.value = v
            return
        trovato = False
        for i, parte in enumerate(v):
            contenuto = parte.text if isinstance(parte, TextBlock) else str(parte)
            if testo in contenuto:
                nuovo = contenuto.replace(testo, token)
                if isinstance(parte, TextBlock):
                    parte.text = nuovo
                else:
                    v[i] = nuovo
                trovato = True
        if trovato:
            cella.value = v
        elif testo in str(v):
            # il valore attraversa piu' run: si appiattisce solo questa cella
            cella.value = str(v).replace(testo, token)
        return

    corrente = _fmt(v)
    if testo == "__APPEND__":
        cella.value = corrente + " " + token if corrente else token
    elif testo in corrente:
        cella.value = corrente.replace(testo, token)



def costruisci_template(prop: Proposta, cartella_templates: str | Path,
                        pattern_nome_file: str | None = None) -> Path:
    """Applica la proposta: scrive template.xlsx tokenizzato + config.json.

    Ritorna la cartella del nuovo PN.
    """
    if not prop.pn:
        raise ValueError("PN mancante nella proposta.")
    dest = Path(cartella_templates) / prop.pn
    dest.mkdir(parents=True, exist_ok=True)

    # rich_text=True: i testi con formattazioni parziali (es. rosso su parte
    # della frase) vengono preservati nel template e nei cicli generati.
    wb = openpyxl.load_workbook(prop.sorgente, rich_text=True)
    ws = wb.active
    merges = _mappa_merge(ws)

    campi_config = []
    for campo in prop.campi:
        if not campo.incluso:
            continue
        # tokenizza la cella principale
        if campo.cella:
            from openpyxl.utils.cell import coordinate_to_tuple
            r, c = coordinate_to_tuple(campo.cella)
            ar, ac = _anchor(ws, r, c, merges)
            ws.cell(row=ar, column=ac).value = campo.token
        # applica le sostituzioni testuali (preservando i run colorati)
        for coord, testo in campo.sostituzioni:
            from openpyxl.utils.cell import coordinate_to_tuple
            r, c = coordinate_to_tuple(coord)
            ar, ac = _anchor(ws, r, c, merges)
            cella = ws.cell(row=ar, column=ac)
            _sostituisci_testo_cella(cella, testo, campo.token)
        campi_config.append({
            "token": campo.token,
            "etichetta": campo.etichetta,
            "tipo": campo.tipo,
            "obbligatorio": campo.obbligatorio,
            "default": campo.default,
            "gruppo": campo.gruppo,
            "aiuto": campo.aiuto,
            "fisso": campo.fisso,
        })

    wb.save(dest / "template.xlsx")
    from motore.generatore import ripara_rich_text_salvato
    ripara_rich_text_salvato(dest / "template.xlsx")

    config = {
        "pn": prop.pn,
        "descrizione": prop.descrizione,
        "template": "template.xlsx",
        "pattern_nome_file": pattern_nome_file
            or "{ORDINE_INTERNO}_ODL_OR_{PN}_REV.{REV}.xlsx",
        "campi": campi_config,
    }
    (dest / "config.json").write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return dest
