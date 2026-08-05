"""Motore di generazione dei RIQ (Rapporto Ispezione Qualita').

Un RIQ e' associato 1:1 a un ODL tramite Ordine Interno + PN: stesso
approccio a token del ciclo di lavoro (``generatore.py``), piu' una
tabella misure a righe/colonne fisse (rilevate dal wizard, vedi
``rileva_riq.py``) e un blocco esito (data, conforme/non conforme,
firma CQ) sincronizzato con l'ODL.

Struttura tipica di un RIQ (famiglia "Mod 02-01_riq_CQ"):
    - testata a token (RIQ, Protocollo, Commessa, Ordine Acquisto,
      Descrizione, Ordine di lavoro, Lotto, PN, DWG, Quantita');
    - tabella "ISPEZIONI DA ESEGUIRE": righe fisse con Posizione,
      Caratteristica, Modalita' di verifica, Strumento, Accettato,
      Scartato, Risultato rilevato (quest'ultimo e' il valore letto
      dal calibro, con eventuale varianza casuale);
    - riga esito: Data, Stato (CONFORME/NON CONFORME), firma CQ;
    - RNC / Richiesta Concessione se non conforme;
    - Redatto da / Approvato da (firma standard, come nell'ODL).
"""

from __future__ import annotations

import json
import random
import re
from dataclasses import dataclass, field as dc_field
from pathlib import Path
from typing import Any

import openpyxl

from .generatore import (
    Campo,
    _parse_data,
    _sostituisci_in_cella,
    _sostituisci_in_rich_text,
    _valore_stringa,
    inserisci_immagine,
    ripara_rich_text_salvato,
    scrivi_data_cella,
)

CONFORME = "CONFORME"
NON_CONFORME = "NON CONFORME"

# Modalita' di firma CQ.
FIRMA_CQ_PNG = "png"                 # solo immagine
FIRMA_CQ_PNG_INIZIALI = "png_iniziali"  # immagine + iniziali/nome accanto
FIRMA_CQ_INIZIALI = "iniziali"       # solo testo (nome/sigla)


# --------------------------------------------------------------------------- #
# Configurazione RIQ
# --------------------------------------------------------------------------- #
@dataclass
class RigaMisura:
    """Una riga della tabella "ISPEZIONI DA ESEGUIRE"."""
    riga: int                      # riga nel foglio Excel
    posizione: str = ""            # numero progressivo caratteristica
    caratteristica: str = ""       # descrizione / nominale + tolleranza
    modalita_verifica: str = ""
    strumento: str = ""
    accettato: str = ""
    scartato: str = ""
    risultato_rilevato: str = ""   # valore misurato (da calibro)
    nominale: float | None = None  # se estraibile da 'caratteristica', per varianza
    tolleranza: float | None = None
    richiede_misura: bool = False  # True se lo strumento e' un calibro (valore numerico atteso)


@dataclass
class ColonneTabellaMisure:
    posizione: str = "A"
    caratteristica: str = "B"
    modalita_verifica: str = "F"
    strumento: str = "G"
    accettato: str = "H"
    scartato: str = "I"
    risultato: str = "J"


@dataclass
class BloccoEsito:
    riga: int
    colonna_data: str
    colonna_stato: str
    cella_firma_cq: str            # cella (o zona unita) per immagine/testo firma CQ


@dataclass
class BloccoNonConformita:
    riga: int
    cella_rnc: str
    cella_concessione: str


@dataclass
class BloccoRedattoApprovato:
    cella_redatto: str | None = None
    cella_approvato: str | None = None


@dataclass
class ConfigRIQ:
    pn: str
    descrizione: str
    template: str
    campi: list[Campo]
    riga_inizio_misure: int
    riga_fine_misure: int
    colonne_misure: ColonneTabellaMisure
    blocco_esito: BloccoEsito
    blocco_non_conformita: BloccoNonConformita | None = None
    blocco_redatto_approvato: BloccoRedattoApprovato | None = None
    pattern_nome_file: str = "{ORDINE_INTERNO}_RIQ_OR_{PN}.xlsx"
    cartella: Path | None = dc_field(default=None)

    @classmethod
    def da_file(cls, percorso: str | Path) -> "ConfigRIQ":
        percorso = Path(percorso)
        dati = json.loads(percorso.read_text(encoding="utf-8"))
        campi = [Campo(**c) for c in dati.get("campi", [])]
        colonne = ColonneTabellaMisure(**dati["colonne_misure"])
        esito = BloccoEsito(**dati["blocco_esito"])
        nc = dati.get("blocco_non_conformita")
        ra = dati.get("blocco_redatto_approvato")
        return cls(
            pn=dati["pn"],
            descrizione=dati.get("descrizione", ""),
            template=dati["template"],
            campi=campi,
            riga_inizio_misure=dati["riga_inizio_misure"],
            riga_fine_misure=dati["riga_fine_misure"],
            colonne_misure=colonne,
            blocco_esito=esito,
            blocco_non_conformita=BloccoNonConformita(**nc) if nc else None,
            blocco_redatto_approvato=BloccoRedattoApprovato(**ra) if ra else None,
            pattern_nome_file=dati.get(
                "pattern_nome_file", "{ORDINE_INTERNO}_RIQ_OR_{PN}_REV.{REV}.xlsx"),
            cartella=percorso.parent,
        )

    def percorso_template(self) -> Path:
        base = self.cartella or Path(".")
        return (base / self.template).resolve()

    def mappa_token(self) -> dict[str, Campo]:
        return {c.token: c for c in self.campi}

    def campo_per_tipo(self, tipo: str) -> Campo | None:
        for c in self.campi:
            if c.tipo == tipo:
                return c
        return None


def percorso_config_riq(cartella_pn: Path) -> Path:
    return cartella_pn / "config_riq.json"


def esiste_riq(cartella_pn: Path) -> bool:
    return percorso_config_riq(cartella_pn).is_file()


# --------------------------------------------------------------------------- #
# Lettura righe misura esistenti dal template (per lo strumento "calibro")
# --------------------------------------------------------------------------- #
_RE_NOMINALE = re.compile(
    r"(?P<nom>-?\d+(?:[.,]\d+)?)\s*\(?\s*[±±]\s*(?P<tol>\d+(?:[.,]\d+)?)"
)


def _estrai_nominale_tolleranza(testo: str) -> tuple[float | None, float | None]:
    """Da un testo tipo '49 (± 0,3) mm' estrae (49.0, 0.3)."""
    if not testo:
        return None, None
    m = _RE_NOMINALE.search(testo)
    if not m:
        return None, None
    try:
        nom = float(m.group("nom").replace(",", "."))
        tol = float(m.group("tol").replace(",", "."))
        return nom, tol
    except ValueError:
        return None, None


def leggi_righe_misura(config: ConfigRIQ) -> list[RigaMisura]:
    """Legge dal template le righe della tabella misure (nominali, strumenti)
    cosi' come compilate nel ciclo campione, per proporle nello strumento di
    inserimento/approvazione valori."""
    wb = openpyxl.load_workbook(config.percorso_template(), data_only=True)
    ws = wb.active
    col = config.colonne_misure
    righe: list[RigaMisura] = []
    for r in range(config.riga_inizio_misure, config.riga_fine_misure + 1):
        posizione = ws[f"{col.posizione}{r}"].value
        caratteristica = ws[f"{col.caratteristica}{r}"].value
        if posizione is None and caratteristica is None:
            continue
        strumento = ws[f"{col.strumento}{r}"].value or ""
        nominale, tolleranza = _estrai_nominale_tolleranza(str(caratteristica or ""))
        righe.append(RigaMisura(
            riga=r,
            posizione=str(posizione or "").strip(),
            caratteristica=str(caratteristica or "").strip(),
            modalita_verifica=str(ws[f"{col.modalita_verifica}{r}"].value or "").strip(),
            strumento=str(strumento).strip(),
            accettato=str(ws[f"{col.accettato}{r}"].value or "").strip(),
            scartato=str(ws[f"{col.scartato}{r}"].value or "").strip(),
            risultato_rilevato="",
            nominale=nominale,
            tolleranza=tolleranza,
            richiede_misura=nominale is not None,
        ))
    return righe


# --------------------------------------------------------------------------- #
# Strumento "calibro": varianza casuale sui valori misurati
# --------------------------------------------------------------------------- #
def applica_varianza(nominale: float, varianza: float) -> float:
    """Genera un valore plausibile nell'intervallo [nominale-varianza,
    nominale+varianza], con distribuzione uniforme."""
    return round(nominale + random.uniform(-varianza, varianza), 3)


def formatta_misura(valore: float, unita: str = "mm") -> str:
    testo = f"{valore:.2f}".replace(".", ",")
    return f"{testo} {unita}".strip()


def genera_valori_calibro(righe: list[RigaMisura], varianza: float) -> None:
    """Compila in-place ``risultato_rilevato`` per le righe che richiedono
    una misura (hanno un nominale numerico), applicando una varianza
    casuale e determinando Accettato/Scartato in base alla tolleranza."""
    for r in righe:
        if not r.richiede_misura or r.nominale is None:
            continue
        tol = r.tolleranza if r.tolleranza is not None else varianza
        valore = applica_varianza(r.nominale, min(varianza, tol) if tol else varianza)
        r.risultato_rilevato = formatta_misura(valore)
        if r.tolleranza is not None and abs(valore - r.nominale) > r.tolleranza:
            r.accettato, r.scartato = "", "X"
        else:
            r.accettato, r.scartato = "V", ""


# --------------------------------------------------------------------------- #
# Firma CQ
# --------------------------------------------------------------------------- #
def applica_firma_cq(ws, cella: str, nome_o_sigla: str, percorso_png: str | Path | None,
                     modo: str) -> None:
    """Scrive la firma CQ nella cella/zona indicata secondo la modalita':
        - FIRMA_CQ_PNG: solo immagine;
        - FIRMA_CQ_PNG_INIZIALI: immagine + nome/sigla in coda al testo esistente;
        - FIRMA_CQ_INIZIALI: solo testo (nome/sigla), nessuna immagine.
    """
    if modo == FIRMA_CQ_INIZIALI or percorso_png is None:
        if ws[cella].value:
            ws[cella] = f"{ws[cella].value} - {nome_o_sigla}"
        else:
            ws[cella] = nome_o_sigla
        return
    inserisci_immagine(ws, cella, str(percorso_png))
    if modo == FIRMA_CQ_PNG_INIZIALI:
        # Le iniziali vanno accanto (stessa cella non e' possibile con
        # un'immagine ancorata: si scrive nella cella immediatamente
        # sottostante alla zona, se libera).
        from openpyxl.utils.cell import coordinate_to_tuple
        from .generatore import _zona_unita
        r_min, c_min, r_max, c_max = _zona_unita(ws, cella)
        cella_sotto = ws.cell(row=r_max + 1, column=c_min)
        if cella_sotto.value in (None, ""):
            cella_sotto.value = nome_o_sigla


# --------------------------------------------------------------------------- #
# Generazione RIQ
# --------------------------------------------------------------------------- #
def _sanitizza_nome(nome: str) -> str:
    from .generatore import _CARATTERI_VIETATI_FILE
    return _CARATTERI_VIETATI_FILE.sub("-", nome).strip()


def nome_file_riq(config: ConfigRIQ, valori: dict[str, Any]) -> str:
    contesto = {"PN": config.pn}
    for c in config.campi:
        chiave = c.token.strip("{}")
        contesto[chiave] = _valore_stringa(c, valori.get(c.token))

    def sostituisci(m: re.Match) -> str:
        return str(contesto.get(m.group(1), m.group(0)))

    nome = re.sub(r"\{([A-Za-z0-9_]+)\}", sostituisci, config.pattern_nome_file)
    return _sanitizza_nome(nome)


def genera_riq(
    config: ConfigRIQ,
    valori: dict[str, Any],
    percorso_out: str | Path,
    righe_misura: list[RigaMisura] | None = None,
    data_esito: Any = None,
    stato: str = CONFORME,
    firma_cq_nome: str | None = None,
    firma_cq_png: str | Path | None = None,
    firma_cq_modo: str = FIRMA_CQ_PNG,
    rnc: str = "",
    concessione: str = "",
    firma_redatto_testo: str | None = None,
    firma_redatto_immagine: str | Path | None = None,
) -> Path:
    """Genera un singolo RIQ .xlsx a partire dal template e dai valori.

    ``data_esito`` deve corrispondere alla data del campo RIQ nell'ODL
    (sincronizzazione richiesta): va passata dal chiamante, non generata
    qui, cosi' ODL e RIQ condividono la stessa fonte del dato.
    """
    from openpyxl.cell.rich_text import CellRichText

    mappa = config.mappa_token()
    wb = openpyxl.load_workbook(config.percorso_template(), rich_text=True)
    for ws in wb.worksheets:
        for riga in ws.iter_rows():
            for cella in riga:
                v = cella.value
                if isinstance(v, CellRichText):
                    if "{{" in str(v):
                        _sostituisci_in_rich_text(cella, v, mappa, valori)
                elif isinstance(v, str) and "{{" in v:
                    _sostituisci_in_cella(cella, v, mappa, valori)

    ws0 = wb.worksheets[0]
    col = config.colonne_misure
    if righe_misura:
        for rm in righe_misura:
            if rm.risultato_rilevato:
                ws0[f"{col.risultato}{rm.riga}"] = rm.risultato_rilevato
            if rm.accettato:
                ws0[f"{col.accettato}{rm.riga}"] = rm.accettato
            if rm.scartato:
                ws0[f"{col.scartato}{rm.riga}"] = rm.scartato

    esito = config.blocco_esito
    if data_esito is not None:
        scrivi_data_cella(ws0, f"{esito.colonna_data}{esito.riga}", data_esito)
    ws0[f"{esito.colonna_stato}{esito.riga}"] = stato
    if firma_cq_nome or firma_cq_png:
        applica_firma_cq(ws0, esito.cella_firma_cq, firma_cq_nome or "",
                         firma_cq_png, firma_cq_modo)

    if stato == NON_CONFORME and config.blocco_non_conformita:
        nc = config.blocco_non_conformita
        if rnc:
            ws0[nc.cella_rnc] = rnc
        if concessione:
            ws0[nc.cella_concessione] = concessione

    if config.blocco_redatto_approvato and config.blocco_redatto_approvato.cella_redatto:
        cella_r = config.blocco_redatto_approvato.cella_redatto
        if firma_redatto_immagine:
            inserisci_immagine(ws0, cella_r, str(firma_redatto_immagine))
        elif firma_redatto_testo:
            corrente = ws0[cella_r].value
            ws0[cella_r] = (f"{corrente} {firma_redatto_testo}" if corrente
                            else firma_redatto_testo)

    percorso_out = Path(percorso_out)
    percorso_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(percorso_out)
    ripara_rich_text_salvato(percorso_out)
    return percorso_out


def valida_riq(config: ConfigRIQ, valori: dict[str, Any]) -> list[str]:
    from .generatore import valida as _valida_generica
    # riusa la stessa logica di validazione campi (obbligatori/tipo)
    class _ConfigCompat:
        campi = config.campi
    return _valida_generica(_ConfigCompat(), valori)
