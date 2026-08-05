"""Test end-to-end del motore di generazione (Fase 1)."""

from __future__ import annotations

import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

from motore import generatore as G
from motore import ordine_interno as oi
from motore import riq as RIQ

RADICE = Path(__file__).resolve().parent
CONFIG = RADICE / "templates" / "70009309-P002" / "config.json"
CONFIG_RIQ = RADICE / "templates" / "55631490" / "config_riq.json"
OUT = RADICE / "output" / "_test"


def main() -> None:
    # 1) incremento ordine interno
    assert oi.serie("6935-26", 4) == ["6935-26", "6936-26", "6937-26", "6938-26"]
    assert oi.incrementa("00042-25") == "00043-25"
    print("[ok] incremento ordine interno")

    config = G.ConfigPN.da_file(CONFIG)

    valori = {
        "{{COMMESSA}}": "SPRBHAL1-875",
        "{{ORDINE_ACQUISTO}}": "2026-OF-0000675",
        "{{ORDINE_INTERNO}}": "6935-26",
        "{{NR_PEZZI}}": "6",
        "{{DATA_ODA}}": "11/05/2026",
        "{{DATA_OI}}": "11/05/2026",
        "{{DATA_CONSEGNA}}": "",           # opzionale, lasciato vuoto
        "{{REV}}": "1",
        "{{LOTTO_MPL-076}}": "DTF066535",
        "{{LOTTO_3V000107}}": "211",       # deve propagarsi in AD24 e S45
        "{{LOTTO_ADESIVO_EPOX}}": "6985446",
        "{{SCADENZA_ADESIVO_EPOX}}": "31/12/2026",
    }

    errori = G.valida(config, valori)
    assert not errori, f"validazione fallita: {errori}"
    print("[ok] validazione input")

    # 2) genera batch di 4, con ultimo ciclo a 2 pezzi
    prodotti = G.genera_batch(
        config, valori, quantita=4, cartella_out=OUT, pezzi_override={3: 2}
    )
    assert len(prodotti) == 4
    nomi = [p.name for p in prodotti]
    print("[ok] generati", len(prodotti), "cicli:")
    for n in nomi:
        print("      -", n)

    attesi = [
        "6935-26_ODL_OR_70009309-P002_REV.1.xlsx",
        "6936-26_ODL_OR_70009309-P002_REV.1.xlsx",
        "6937-26_ODL_OR_70009309-P002_REV.1.xlsx",
        "6938-26_ODL_OR_70009309-P002_REV.1.xlsx",
    ]
    assert nomi == attesi, f"nomi file inattesi: {nomi}"
    print("[ok] nomi file e incremento ordine interno")

    # 3) verifica contenuto del primo e dell'ultimo ciclo
    def celle_testo(ws):
        out = {}
        for riga in ws.iter_rows():
            for c in riga:
                if c.value is not None:
                    out[c.coordinate] = c.value
        return out

    wb0 = openpyxl.load_workbook(prodotti[0])
    ws0 = wb0.active
    testo = celle_testo(ws0)

    # nessun token residuo
    residui = [f"{k}={v}" for k, v in testo.items()
               if isinstance(v, str) and "{{" in v]
    assert not residui, f"token non sostituiti: {residui}"
    print("[ok] nessun token residuo")

    # valori attesi
    assert ws0["C5"].value == "SPRBHAL1-875"
    assert ws0["V5"].value == "6935-26"
    assert ws0["E7"].value == 6 and isinstance(ws0["E7"].value, int)
    assert isinstance(ws0["M7"].value, datetime), f"DATA_ODA non e' datetime: {ws0['M7'].value!r}"
    assert ws0["M7"].value == datetime(2026, 5, 11)
    assert ws0["AF7"].value is None  # data consegna vuota
    print("[ok] valori testata (testo/int/data tipizzati)")

    # propagazione lotto 211: AD24 e testo fase 60 (S45)
    assert str(ws0["AD24"].value) == "211"
    assert ws0["S45"].value == "Lotto: 211", f"S45={ws0['S45'].value!r}"
    assert ws0["S40"].value == "Lotto: 6985446", f"S40={ws0['S40'].value!r}"
    assert ws0["AA40"].value == "Scadenza: 31/12/2026", f"AA40={ws0['AA40'].value!r}"
    print("[ok] propagazione lotti e sostituzione dentro le fasi")

    # inserti lotti lasciati vuoti -> celle vuote (nessun token)
    assert ws0["AD28"].value is None
    print("[ok] lotti opzionali vuoti -> celle vuote")

    # ultimo ciclo: ordine interno + pezzi override
    wbL = openpyxl.load_workbook(prodotti[3])
    wsL = wbL.active
    assert wsL["V5"].value == "6938-26"
    assert wsL["E7"].value == 2, f"override pezzi ultimo ciclo fallito: {wsL['E7'].value}"
    print("[ok] override pezzi ultimo ciclo (=2) e ordine interno finale")

    # 4) fedelta': logo/immagini e impostazioni di stampa preservate
    z = zipfile.ZipFile(prodotti[0])
    nomi_zip = z.namelist()
    assert any("media" in n for n in nomi_zip), "immagini perse!"
    assert ws0.page_setup.orientation == "landscape"
    assert ws0.page_setup.scale == 78
    assert ws0.print_area == "'Ciclo di lavoro'!$A$1:$AP$55"
    print("[ok] logo/immagini + stampa (landscape, scala 78, area A1:AP55) preservati")

    print("\nTUTTI I TEST PASSATI. File in:", OUT)


def test_riq() -> None:
    config_riq = RIQ.ConfigRIQ.da_file(CONFIG_RIQ)

    # 1) lettura righe misura dal template + varianza casuale
    righe = RIQ.leggi_righe_misura(config_riq)
    assert righe, "nessuna riga misura letta dal template RIQ"
    n_con_nominale = [r for r in righe if r.richiede_misura]
    assert n_con_nominale, "nessuna riga con nominale/tolleranza riconosciuti"
    RIQ.genera_valori_calibro(righe, varianza=0.05)
    for r in n_con_nominale:
        assert r.risultato_rilevato, f"riga {r.riga} senza risultato dopo varianza"
        valore = float(r.risultato_rilevato.split()[0].replace(",", "."))
        assert abs(valore - r.nominale) <= max(r.tolleranza, 0.05) + 1e-9, (
            f"riga {r.riga}: valore {valore} fuori range plausibile")
    print("[ok] riq: lettura righe misura + varianza casuale")

    # 2) validazione + generazione con link 1:1 ODL<->RIQ (stesso Ordine Interno)
    valori = {
        "{{RIQ}}": "PN-55631490 Rev.0",
        "{{PROTOCOLLO}}": "2026-321",
        "{{CLIENTE}}": "Avio spa",
        "{{COMMESSA}}": "Aster",
        "{{ORDINE_ACQUISTO}}": "C120790",
        "{{DESCRIZIONE}}": "Honeycomb N.3 Aletta Mobile",
        "{{ORDINE_INTERNO}}": "7000-26",   # deve coincidere con l'ODL corrispondente
        "{{LOTTO}}": "150124",
        "{{PN}}": "55631490",
        "{{DWG}}": "AST51000450-F",
        "{{QUANTITA_PRODOTTE}}": "5",
        "{{PERC_CAMPIONAMENTO}}": "20%",
    }
    errori = RIQ.valida_riq(config_riq, valori)
    assert not errori, f"validazione RIQ fallita: {errori}"
    print("[ok] riq: validazione input")

    out = RIQ.genera_riq(
        config_riq, valori, OUT / "riq" / "7000-26_RIQ_TEST.xlsx",
        righe_misura=righe, data_esito=datetime(2026, 7, 27),
        stato=RIQ.CONFORME, firma_cq_nome="Mario Rossi (MRO)",
        firma_cq_png=None, firma_cq_modo=RIQ.FIRMA_CQ_INIZIALI,
    )
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws["G7"].value == "7000-26", "Ordine Interno RIQ non scritto correttamente"
    assert ws["B8"].value == "55631490"
    assert ws["A31"].value == "27/07/2026"
    assert ws["D31"].value == RIQ.CONFORME
    assert "Mario Rossi" in str(ws["H31"].value)
    residui = [c.coordinate for row in ws.iter_rows() for c in row
               if isinstance(c.value, str) and "{{" in c.value]
    assert not residui, f"token RIQ non sostituiti: {residui}"
    print("[ok] riq: generazione, sincronizzazione data/OI, firma CQ, nessun token residuo")

    # 3) esito non conforme con RNC/Concessione
    out_nc = RIQ.genera_riq(
        config_riq, valori, OUT / "riq" / "7001-26_RIQ_NC_TEST.xlsx",
        righe_misura=righe, data_esito=datetime(2026, 7, 27),
        stato=RIQ.NON_CONFORME, firma_cq_nome="Mario Rossi (MRO)",
        firma_cq_png=None, firma_cq_modo=RIQ.FIRMA_CQ_INIZIALI,
        rnc="RNC-2026-05", concessione="RC-2026-02",
    )
    wsn = openpyxl.load_workbook(out_nc).active
    assert wsn["D31"].value == RIQ.NON_CONFORME
    assert wsn[config_riq.blocco_non_conformita.cella_rnc].value == "RNC-2026-05"
    assert wsn[config_riq.blocco_non_conformita.cella_concessione].value == "RC-2026-02"
    print("[ok] riq: esito non conforme con RNC/Concessione")

    print("\nTUTTI I TEST RIQ PASSATI. File in:", OUT / "riq")


if __name__ == "__main__":
    main()
    test_riq()
